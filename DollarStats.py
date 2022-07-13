import requests as rq
import time
from openpyxl import Workbook


wd = Workbook()
ws = wd.active
minute_start = int(time.strftime('%M'))
high_value = [0]
running = True

while running:
    # Starts at the new minute
    if int(time.strftime('%S')) == 0:
        while running:
            minute = int(time.strftime('%M'))
            # If pass 1 minute after starting the program stops
            if minute - minute_start >= 2:
                print('stop')
                running = False
            else:
                # https://docs.awesomeapi.com.br/api-de-moedas
                dic = rq.get('https://economia.awesomeapi.com.br/last/USD-BRL,EUR-BRL,BTC-BRL').json()

                data = [['Time current:', time.ctime()],
                    ['Dollar Now:', dic['USDBRL']['bid']],
                    ['Dollar Highest price:', dic['USDBRL']['high']],
                    ['Dollar Lowest price:', dic['USDBRL']['low']],
                ]

                for row in data:
                    ws.append(row)
                    high_value.append(float(dic['USDBRL']['bid']))
                time.sleep(19)

ws['D1'] = 'Max value:'
ws['E1'] = max(high_value)

# Save in Excel document
wd.save('DollarStats.xlsx')
